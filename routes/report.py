from flask import render_template,redirect,request,url_for,Blueprint,send_file
from models import Item, ItemHistory, User, Personnel
from extensions import db
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import pandas as pd
from .login import admin_required
import io

report_bp = Blueprint("report_bp",__name__,url_prefix="/report")



@report_bp.route('/export_report')
def export_report():
    # مشخص کردن جدول‌ها و نام شیت‌ها
    models = [
        (Item, "Items"),
        (ItemHistory, "ItemHistory"),  # شیت جدا برای ItemHistory
        (User, "Users"),
        (Personnel, "Personnel"),
    ]

    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for model, sheet_name in models:
            records = model.query.all()
            if not records:
                continue

            # استخراج داینامیک ستون‌ها
            data = [{col.name: getattr(r, col.name) for col in r.__table__.columns} for r in records]
            df = pd.DataFrame(data)
            df.to_excel(writer, index=False, sheet_name=sheet_name)
    output.seek(0)

    # اعمال فرمت‌ها با openpyxl
    wb = load_workbook(output)
    for ws in wb.worksheets:
        ws.sheet_view.rightToLeft = True
        center_alignment = Alignment(horizontal="center", vertical="center")
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = center_alignment
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 3

    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    return send_file(final_output,
                     as_attachment=True,
                     download_name="all_tables_report.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")







@report_bp.route("/report")
def report():
    args = request.args

    field_mapping = {
        "property_code": Item.property_code,
        "project_code": Item.project_code,
        "warehouse_location": Item.warehouse_location,
        "row": Item.row,
        "user": Item.user,
        "company": Item.company,
        "category": Item.category,
        "personnel_code": Item.personnel_code,
        "current_location": Item.current_location,
        "system_identification_code": Item.system_identification_code,
        "model": Item.model,
        "serial_number": Item.serial_number,
        "recipient_delivery": Item.recipient_delivery,
        "closed": Item.closed
    }

    query = Item.query
    has_filter = False

    # ---- اعمال فیلترهای فرم ----
    for arg_key, model_field in field_mapping.items():
        value = args.get(arg_key)
        if value:
            has_filter = True
            query = query.filter(model_field.ilike(f"%{value}%"))

    if not has_filter:
        return render_template("report_panel/report.html", 
                               items=[],
                               group_data=None,
                               group_field=None)

    # ---- گروه‌بندی ----
    group_field = args.get("group_field")
    group_data = None

    if group_field in field_mapping:
        column = field_mapping[group_field]
        group_data = (
            query.with_entities(column.label("field"), db.func.count().label("cnt"))
                 .group_by(column)
                 .order_by(db.func.count().desc())
                 .all()
        )

    # ---- نتایج لیستی ----
    items = query.order_by(Item.property_code.desc()).all()

    return render_template("report_panel/report.html",
                           items=items,
                           group_data=group_data,
                           group_field=group_field)




@report_bp.route("/suggest", methods=["GET"])
def suggest():
    args = request.args

    field = args.get("field")
    value = args.get("value", "")

    field_mapping = {
       "property_code": Item.property_code,
        "project_code": Item.project_code,
        "warehouse_location":Item.warehouse_location,
        "row":Item.row,
        "user": Item.user,
        "company": Item.company,
        "category": Item.category,
        "personnel_code":Item.personnel_code,
        "current_location":Item.current_location,
        "system_identification_code":Item.system_identification_code,
        "model":Item.model,
        "serial_number":Item.serial_number,
        "recipient_delivery":Item.recipient_delivery,
        "closed":Item.closed
    }

    if field not in field_mapping:
        return {"error": "invalid field"}, 400

    query = Item.query

    # اعمال فیلتر برای فیلدهای دیگر
    for key, column in field_mapping.items():
        if key != field:
            v = args.get(key)
            if v:
                query = query.filter(column.ilike(f"%{v}%"))

    # دریافت همه مقادیر
    suggestions = (
        query.with_entities(field_mapping[field])
        .distinct()
        .filter(field_mapping[field].ilike(f"%{value}%"))
        .order_by(field_mapping[field])
        
        .all()
    )

    return {"suggestions": [s[0] for s in suggestions if s[0]]}